"""
@Time : 2024/4/17 0017 21:07
@Auth : Davinstein
@File : ChatPDFScore.py
"""
import threading
import queue
import re
import os
import time
from utils import upload_chatpdf_file, query_chatpdf


BATCH_SIZE = 20
class Student_work:
    # zzh
    # Chatpdf_api_key = 'sec_rD2WAd1bO7ELLJ0lNMBPkANi7Ux2ndN9'
    # zwl
    # Chatpdf_api_key = 'sec_gUKCF92a9fGRIltq4vCWUWgqbtqwuryQ'
    #yj
    Chatpdf_api_key = 'sec_7qcVWTEz4yjYjvA8ZU0X86cxhqjymJZO'
    Prompt = '你现在是一名网络安全实践课程的助教，请使用不变的评分标准为刚才的课程报告pdf进行打分，满分100分，并在回复中使用数字给出最后得分，并且不要出现其他分数，可以有一些评价打分依据。'
    def __init__(self, path):
        self.root_path = path
        self.result_queue = queue.Queue()
        self.pdf_file_num = self.Get_pdf_num()
        self.score_info_list = []
    def Get_pdf_num(self):
        file_list = os.listdir(self.root_path)
        count = 0
        for file in file_list:
            if file.endswith('.pdf'):
                count += 1
        return count

    def Get_Score_Single_PDF(self, file_name):
        print('-------------create single file thread:-------------------', file_name)
        file_path = os.path.join(self.root_path, file_name)
        paper_source_id = upload_chatpdf_file(file_path, Student_work.Chatpdf_api_key)
        # query_chatpdf(Student_work.chatpdf_api_key, paper_source_id, Student_work.Prompt)
        single_thread = threading.Thread(target=Get_Score_from_response, args=(Student_work.Chatpdf_api_key, file_name, paper_source_id,
                                                                    Student_work.Prompt, self.result_queue))
        return single_thread
    def Get_Score_Queue_All_PDF(self):
        all_threads = [self.Get_Score_Single_PDF(file_name) for file_name in os.listdir(self.root_path) if file_name.endswith('.pdf')]
        if len(all_threads) == 0:
            self.score_info_list = []
        # for file_name in os.listdir(self.root_path):
        #     self.Get_Score_Single_PDF(file_name)
        for t in all_threads:
            t.start()
            time.sleep(5)
        results = [self.result_queue.get() for i in range(self.pdf_file_num)]
        self.score_info_list = results
        print(results)
    def Save_single_file_score_info(self, file_name, score, response):
        tup = (file_name, score, response)
        self.score_info_list.append(tup)
        return tup
    def Get_Student_Score_Info(self):
        return self.score_info_list


def Get_Score_from_response(chatpdf_api_key, file_name, paper_source_id, prompt, result_queue):
    '''
    多线程处理根据source_id得到对应分数
    Args:
        response:
        source_id:
        result_queue:

    Returns:

    '''
    print(f'-----------------Thread {paper_source_id} starting-----------------')
    response = query_chatpdf(chatpdf_api_key, paper_source_id, prompt)
    Wrong_score = -1
    if response:
        number_list = [int(number) for number in re.findall(r'\d+', response)]
    # print(number_list)

        if number_list:
            for number in number_list:
                if number >=50 and number <= 100:
                    tup = (file_name, number, response)
                    result_queue.put(tup)
                    return tup
        else:
            print('No scores included in the response...')
            tup = (file_name, Wrong_score, response)
            result_queue.put(tup)
            return tup
    else:
        result_queue.put((file_name, Wrong_score, response))
        return (file_name, Wrong_score, response)

import zipfile
import os
import re
WRONG_NAME = "WRONG-NAME"
WRONG_ID = "WRONG-ID"
def extract_name_id_from_file(zip_name):
    # match = re.search(r'(\d{9})([\u4e00-\u9fa5]+)', zip_name)
    match = re.search(r'(\d{9})\W*\s*([\u4e00-\u9fa5]+)', zip_name)
    if match:
        student_id = match.group(1)  # 学号部分
        student_name = match.group(2)  # 姓名部分

        # print(f"学号: {student_id}")


        # print(f"姓名: {student_name}")
        return student_id, student_name
    else:
        print(f"{zip_name} 未能匹配到学号和姓名")
        return WRONG_ID, WRONG_NAME

def read_and_upload_pdf(lecture_zip_file):
    with zipfile.ZipFile(lecture_zip_file, 'r') as lecture_zip:
        for student_zip_info in lecture_zip.infolist():
            student_zip_filename = student_zip_info.filename
            stu_id, stu_name = extract_name_id_from_file(student_zip_filename)
            print(f"Extracting {stu_name} (ID: {stu_id})...")

            with lecture_zip.open(student_zip_filename) as stu_file:
                pdf_files = [name for name in stu_file.namelist() if name.endwith('.pdf')]
                for pdf_file in pdf_files:
                    with stu_file.open(pdf_file) as pdf_f:
                        pdf_content = pdf_f.read()

from openpyxl import Workbook, load_workbook
def read_first_column(lecture):
    root = os.getcwd()
    file_name_list = os.listdir(root)
    all_column_list = []
    for file in file_name_list:
        if os.path.isfile(os.path.join(root, file)):
            if lecture in file and file.endswith('.xlsx'):
                wb = load_workbook(file)
                # 选择活动工作表，默认是第一个工作表
                ws = wb.active

                # 提取第一列的数据
                first_column = [cell.value for cell in ws['A'] if cell.value is not None]
                all_column_list += first_column
                # 打印第一列的数据
                # print(first_column)

                # 关闭工作簿
                wb.close()
    # print(all_column_list)
    return all_column_list
    # if os.path.exists(lecture):
    #     wb = load_workbook(lecture)
    #     # 选择活动工作表，默认是第一个工作表
    #     ws = wb.active
    #
    #     # 提取第一列的数据
    #     first_column = [cell.value for cell in ws['A'] if cell.value is not None]
    #
    #     # 打印第一列的数据
    #     # print(first_column)
    #
    #     # 关闭工作簿
    #     wb.close()
    #     return first_column
    # else:
    #      return []
from datetime import datetime
def cal_and_output():
    lecture_list = ['第一周', '第二周', '第三周', '第四周', '第五周', '第六周']
    # stu_already_list = ['012130109', '022120131', '052110326', "052110502", '052110525', "062120222", '072180201', '092001319',
    #                     '092101331', '161920320','162020119', '162120101', '162120102', '162120103']
    for lecture in lecture_list:
        print(lecture)
        wb = Workbook()
        ws = wb.active
        stu_list = os.listdir(lecture)

        now_time = datetime.now()
        out_ex = lecture + str(now_time.hour) + str(now_time.minute) +  '.xlsx'
        STU_count = 0
        finished_list = read_first_column(lecture)
        try:
            for stu_work in stu_list:
                id, name = extract_name_id_from_file(stu_work)
                stu_header_info = (id, name)
                if id in finished_list:
                    print(f'#####{id} is already in result.######')
                    continue
                stu_file_list = os.listdir(os.path.join(lecture, stu_work))
                stu_pdf_lsit = [pdf for pdf in stu_file_list if pdf.endswith('.pdf')]
                stu_path = os.path.join(lecture, stu_work)

                stu = Student_work(stu_path)
                stu.Get_Score_Queue_All_PDF()


                for row in stu.score_info_list:
                    # print(row)
                    new_row = stu_header_info + row
                    ws.append(new_row)
                STU_count += 1
                # if STU_count > BATCH_SIZE:
                #     print('大于batch_size....')
                #     break
            wb.save(out_ex)
            print(f"本次统计人数：{STU_count}")
            print(f"Excel 文件 '{out_ex}' 已保存成功.")
        except Exception as e:
            wb.save(out_ex)

            print('发生异常，但已保存当前进度。')
            raise e  # 重新抛出异常以便进一步处理

if __name__ == '__main__':
    cal_and_output()
